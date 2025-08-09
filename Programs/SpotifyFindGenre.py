# Nicolas Valdez Mon Jul 14 12:53:27 2025
# SpotifyFindGenre
# Get genres associated with an artist for a specified track.
# Input(s)
# MusicChannelMaster.xlsx
# Output
# SpotifyMusicGenres.csv
# Main Resources
# https://spotipy.readthedocs.io/en/2.25.1/index.html#spotipy.client.Spotify.artist
# https://developer.spotify.com/dashboard
# https://github.com/spotipy-dev/spotipy/tree/2.22.1

import spotipy
from spotipy.oauth2 import SpotifyClientCredentials
from openpyxl import load_workbook
import time
import csv

# Welcoming Statement
print('Get genres associated with an artist for a specified track.')

# Import workbook and worksheet
file_name='MusicChannelMaster.xlsx'
active_sheet='Spotify Track Only'
wb=load_workbook(file_name)
ws=wb[active_sheet]

# Delimiter
delimiter='%'

# Auth manager
keys=[]
keys_file_name='APIKeys.csv'
with open(keys_file_name,'r',newline='') as in_file:
    reader=csv.reader(in_file,delimiter=delimiter)
    for row in reader:
        keys.append(row)
client_id=keys[1][1]
client_secret=keys[2][1]
auth_manager=SpotifyClientCredentials(client_id,client_secret)

# Select table rows
table=ws.tables
values_only=True
min_row=2
min_col=1
max_col=2
date_url=[]
for row in ws.iter_rows(values_only=values_only,
                        min_row=min_row,
                        min_col=min_col,
                        max_col=max_col):
    row_date=''.join(str(row[0]))
    row_url=''.join(row[1])
    date_url.append((row_date,row_url))

# Initialize spotipy
sp=spotipy.Spotify(auth_manager=auth_manager)

# Call constants
rate_limit=0.3
num_genres=3

# Genres export file
export_file_1='SpotifyMusicGenres.csv'
export_file_2='SpotifyFailedToFindGenres.csv'
col1='Date'
col2='Track'
col3='Artist'
col4='Tags'

# Write headers
with open(export_file_1,'w',encoding='utf-8') as in_file:
    in_file.write(f'{col1}{delimiter}{col2}{delimiter}{col3}{delimiter}{col4}\n')
with open(export_file_2,'w',encoding='utf-8') as in_file:
    in_file.write(f'{col1}{delimiter}{col2}{delimiter}{col3}\n')

# Check for zero genres
def no_genres_error(genres):
    if len(genres)==0:
        raise ZeroDivisionError
        
for info in date_url:
    try:
        track_url=info[1]
        # Make call
        r_track=sp.track(track_url)
        
        # Get track name and artist name
        t=r_track['name']
        c=r_track['artists'][0]['name']
        d=info[0]
        artist_uri=r_track['artists'][0]['uri']
        
        # Get genre
        r_genres=sp.artist(artist_uri)
        genres=r_genres['genres'][0:num_genres]
        no_genres_error(genres)
        print(f'{t} - {c}')
        print(f'{genres}\n')
        
        # Append track and artist to export_file_1
        with open(export_file_1,'a',encoding='utf-8') as in_file:
            in_file.write(f'{d}{delimiter}{t}{delimiter}{c}{delimiter}{genres}\n')
    except ZeroDivisionError:
        with open(export_file_2,'a',encoding='utf-8') as in_file:
            in_file.write(f'{d}{delimiter}{t}{delimiter}{c}\n')
        print(f'No tags exist: {t} - {c}\n')
        pass
    # Rate limit
    time.sleep(rate_limit)

# Ending Note
print('Program Ends')
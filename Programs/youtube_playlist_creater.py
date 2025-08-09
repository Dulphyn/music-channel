import os
import google_auth_oauthlib.flow
import googleapiclient.discovery
import googleapiclient.errors
from openpyxl import load_workbook

# Import workbook and worksheet
file_name='MusicChannelMaster.xlsx'
active_sheet='YT ID No Playlist'
wb=load_workbook(file_name)
ws=wb[active_sheet]

# Select table rows
table=ws.tables
values_only=True
min_row=2
video_ids=[]

for row in ws.iter_rows(values_only=values_only,min_row=min_row):
    row_id=''.join(row[1])
    video_ids.append(row_id)
    
scopes = ["https://www.googleapis.com/auth/youtube.force-ssl"]

# Playlist name
playlist_name='The Music of #Music'

def add_video_to_playlist(youtube, video_id, playlist_id,playlist_name=playlist_name):
    request = youtube.playlistItems().insert(
        part="snippet",
        body={
            "snippet": {
                "playlistId": playlist_id,
                "resourceId": {
                    "kind": "youtube#video",
                    "videoId": video_id
                }
            }
        }
    )
    response = request.execute()
    print(f"Added video {video_id} to playlist {playlist_name}")
    
def main():
    # Disable OAuthlib's HTTPS verification when running locally.
    # DO NOT leave this option enabled in production.
    os.environ["OAUTHLIB_INSECURE_TRANSPORT"] = "1"

    api_service_name = "youtube"
    api_version = "v3"
    client_secrets_file = "client_secrets.json"

    # Get credentials and create an API client
    flow = google_auth_oauthlib.flow.InstalledAppFlow.from_client_secrets_file(
        client_secrets_file, scopes)
    credentials = flow.run_local_server(port=0)
    youtube = googleapiclient.discovery.build(
        api_service_name, api_version, credentials=credentials)

    # Specify your playlist ID here
    playlist_id='PLdbBd6kWRFLagfybT9EHt-1kMgaS1XQq4'
    
    # How many items already in playlist
    progress=652
    
    # Add search results to the playlist
    for video_id in video_ids[progress+1:len(video_ids)]:
        try:
            add_video_to_playlist(youtube, video_id, playlist_id)
        except Exception as e:
            print(e)
            print(f'{video_id} failed.')
            pass
        

    


if __name__ == "__main__":
    main()
    print('Program Ends')

# Nicolas Valdez Wed Jul  2 17:07:32 2025
# GetChannelAndTrack
# Gets channel and title from Youtube links
# Input(s)
# MusicChannelMaster.xlsx
# Output
# TrackAndChannel.csv
# AdditionalInformation.txt
# Main Resources
# https://ytmusicapi.readthedocs.io/en/stable/reference/browsing.html#ytmusicapi.YTMusic.get_song
# https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html#openpyxl.worksheet.worksheet.Worksheet.set_printer_settings
# https://www.geeksforgeeks.org/python/get-values-of-all-rows-in-a-particular-column-in-openpyxl-python/

from ytmusicapi import YTMusic
from openpyxl import load_workbook
import time

# Welcoming Statement
print('Gets channel and title from Youtube links.')

# Set up YTMusic
ytmusic = YTMusic()

# Import workbook and worksheet
file_name='MusicChannelMaster.xlsx'
active_sheet='YT ID No Playlist'
wb=load_workbook(file_name)
ws=wb[active_sheet]

# Select table rows
table=ws.tables
values_only=True
min_row=2
videoIDs=[]

for row in ws.iter_rows(values_only=values_only,min_row=min_row):
    row_date=''.join(str(row[0]))
    row_id=''.join(row[1])
    videoIDs.append((row_date,row_id))

# Get title and channel name
def get_song_info(progress,ID):
            song=ytmusic.get_song(videoId=ID)
            title=song['videoDetails']['title']
            channel=song['videoDetails']['author']
            progress+=1
            return(title,channel,progress)

# Youtube link
date_title_channel=[]
progress=0
failed_ids=[]
rate_limit=0.2 # In seconds
padding=3
ind_start=0
ind_end=len(videoIDs)+1
for info in videoIDs[ind_start:ind_end]:
    try:
        date=info[0]
        ID=info[1]
        song_info=get_song_info(progress,ID)
        title=song_info[0]
        channel=song_info[1]
        progress=song_info[2]
        date_title_channel.append((date,song_info[0],song_info[1]))
        print(f'#{progress}'.ljust(padding),f'| {channel} - {title}')
        time.sleep(rate_limit)
    except KeyError:
        failed_ids.append((date,ID))
        print(f'Failed ID: {ID}')
        pass
    except UnicodeEncodeError:
        failed_ids.append((date,ID))
        print(f'Failed ID: {ID}')
        print('UnicodeEncodeError: Something went wrong.')
        pass
    
# Proportion of titles and channels successfully got
proportion=f'{progress}/{len(videoIDs)}'
print(f'{proportion} converted.')

# Export information
export_file_1='TrackAndChannel.csv'
col1='Date'
col2='Track'
col3='Channel'
delimiter='%'
export_file_2='FailedIDs.csv'
file2_col2='Failed ID'

# Track and channel data
with open(export_file_1,'w',encoding='utf-8') as in_file:
    in_file.write(f'{col1}{delimiter}{col2}{delimiter}{col3}\n')
    for info in date_title_channel:
        d=info[0]
        t=info[1]
        c=info[2]
        in_file.write(f'{d}{delimiter}{t}{delimiter}{c}\n')

# Additional Information
with open(export_file_2,'w',encoding='utf-8') as in_file:
    in_file.write(f'{col1}{delimiter}{file2_col2}\n')
    for info in failed_ids:
        d=info[0]
        ID=info[1]
        in_file.write(f'{d}{delimiter}{ID}\n')
        
# Ending Note
print('Program Ends')
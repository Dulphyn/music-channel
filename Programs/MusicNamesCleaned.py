# Nicolas Valdez Wed Jul  9 13:54:03 2025
# MusicNamesCleaned
# Get proper track name and artist
# Input(s)
# MusicChannelMaster.xlsx
# Output
# CleanedTracks.csv
# FailedTracks.csv
# Main Resources
# https://www.last.fm/api/show/track.search

import requests
import time
from openpyxl import load_workbook
import csv

# Import workbook and worksheet
file_name='MusicChannelMaster.xlsx'
active_sheet='TrackAndChannel'
wb=load_workbook(file_name)
ws=wb[active_sheet]

# Select table rows
table=ws.tables
values_only=True
min_row=2
date_col=1
track_col=2
artist_col=3
dates=[]
track_titles=[]
channel_names=[]
for row in ws.iter_rows(values_only=values_only,
                        min_row=min_row,
                        min_col=date_col,
                        max_col=date_col):
    row_string=''.join(row)
    dates.append(row_string)
    
for row in ws.iter_rows(values_only=values_only,
                        min_row=min_row,
                        min_col=track_col,
                        max_col=track_col):
    row_string=''.join(row)
    track_titles.append(row_string)

for row in ws.iter_rows(values_only=values_only,
                        min_row=min_row,
                        min_col=artist_col,
                        max_col=artist_col):
    row_string=''.join(row)
    channel_names.append(row_string)

date_title_channel=tuple(zip(dates,track_titles,channel_names))

# Delimiter
delimiter='%'

# Auth manager
keys=[]
keys_file_name='APIKeys.csv'
with open(keys_file_name,'r',newline='') as in_file:
    reader=csv.reader(in_file,delimiter=delimiter)
    for row in reader:
        keys.append(row)
        
# Base URL
base_url='https://ws.audioscrobbler.com/2.0/'

# Headers
user_agent='Dataquest'
headers={'user-agent':user_agent}

# Payload constants
api_key=keys[1][0]
method='track.search'
format_='json'

# Set payload variables
def payload_info(title,api_key=api_key,method=method,format_=format_):
    track=title
    payload={
        'api_key':api_key,
        'method':method,
        'format':format_,
        'track':track,
        }
    return(payload)

# Create API request
def get_request(payload,header,base_url=base_url):
    response=requests.get(base_url,headers=header,params=payload)
    return(response)

# Get and print API call status
def API_status(r):
    status=r.status_code
    t=r.json()['results']['trackmatches']['track'][0]['name']
    c=r.json()['results']['trackmatches']['track'][0]['artist']
    print(f'\nStatus Code: {status}')
    print(f'Track: {t}')
    print(f'Artist: {c}')
    return(t,c)
        
def failed_song(failed_track):
    title=failed_track[1]
    failed_tracks.append(failed_track)
    print(f'Failed song: {title}\n')
    
# API call constants
num_tags=3
rate_limit=0.3 # In seconds

# Genres export file
export_file_1='CleanedTracks.csv'
export_file_2='FailedTracks.csv'
col1='Date'
col2='Track'
col3='Artist'
file2_col3='Channel'
successful_tracks=[]
failed_tracks=[]

# Export successful tracks
# def write_successful(export_file=export_file_1,col1=col1,col2=col2,delimiter=delimiter,tracks_and_artists=successful_tracks):
#     with open(export_file,'w',encoding='utf-8') as in_file:
#         in_file.write(f'{col1}{delimiter}{col2}\n')
#         for info in tracks_and_artists:
#             t=info[0]
#             c=info[1]
#             in_file.write(f'{t}{delimiter}{c}\n')
            
# def write_failed(export_file=export_file_2,tracks=failed_tracks,col1=col1):
#     with open(export_file,'w',encoding='utf-8') as in_file:
#         in_file.write(f'{progress}/{total}\n')
#         in_file.write(f'{col1}\n')
#         for track in tracks:
#             in_file.write(f'{track}\n')
    
# Make calls
progress=0
total=len(date_title_channel)
for info in date_title_channel:
    try:
        date=info[0]
        title=info[1]
        channel=info[2]
        # Call last.fm
        payload=payload_info(title)
        r=get_request(payload,headers)
        
        # Call status
        status=API_status(r)
        
        # Add to successful songs list
        successful_track=((date,status[0],status[1]))
        successful_tracks.append(successful_track)
        
        # Increment progress bar
        progress+=1
        print(f'{progress}/{total}')
    except KeyError:
        print('\nKeyError')
        failed_track=(info)
        failed_song(failed_track)
        pass
    except ZeroDivisionError:
        print('\nZeroDivisionError')
        failed_track=(info)
        failed_song(failed_track)
        pass 
    except IndexError:
        print('\nIndexError')
        failed_track=(info)
        failed_song(failed_track)
        pass
    time.sleep(rate_limit)
            
# Write export files
with open(export_file_1,'w',encoding='utf-8') as in_file:
    in_file.write(f'{col1}{delimiter}{col2}{delimiter}{col3}\n')
    for info in successful_tracks:
        d=info[0]
        t=info[1]
        c=info[2]
        in_file.write(f'{d}{delimiter}{t}{delimiter}{c}\n')
        
with open(export_file_2,'w',encoding='utf-8') as in_file:
    in_file.write(f'{col1}{delimiter}{col2}{delimiter}{file2_col3}\n')
    for info in failed_tracks:
        d=info[0]
        t=info[1]
        c=info[2]
        in_file.write(f'{d}{delimiter}{t}{delimiter}{c}\n')

# Ending Note
print('Program Ends')